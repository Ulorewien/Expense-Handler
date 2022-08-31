from openpyxl import *
from openpyxl.styles import PatternFill
from dependencies.specifications import *

def createWorkbook(month):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = month
    worksheet["A1"] = "Date"
    worksheet["B1"] = "Expense"
    worksheet["C1"] = "Amount"
    worksheet["D1"] = "Method"
    worksheet["E1"] = "Funds"
    worksheet["F1"] = "Amount"
    worksheet["G1"] = "Method"
    worksheet["H1"] = "Income"
    worksheet["I1"] = "Amount"
    worksheet["J1"] = "Savings"

    for column in range(1, worksheet.max_column + 1):
        setColor(worksheet.cell(1, column), "headers")
        setColor(worksheet.cell(2, column), "headers")
        setColor(worksheet.cell(3, column), "headers")

    return workbook

def createSheet(workbook, month):
    workbook.create_sheet(month)
    worksheet = workbook[month]
    worksheet.title = month
    worksheet["A1"] = "Date"
    worksheet["B1"] = "Expense"
    worksheet["C1"] = "Amount"
    worksheet["D1"] = "Method"
    worksheet["E1"] = "Funds"
    worksheet["F1"] = "Amount"
    worksheet["G1"] = "Method"
    worksheet["H1"] = "Income"
    worksheet["I1"] = "Amount"
    worksheet["J1"] = "Savings"

    for column in range(1, worksheet.max_column + 1):
        setColor(worksheet.cell(1, column), "headers")
        setColor(worksheet.cell(2, column), "headers")

def setColor(cell, type):
    specs = getSpecifications()
    cell.fill = PatternFill("solid", start_color = specs["Excel"][type])

def setRowColor(worksheet, row):
    setColor(worksheet.cell(row, 1), "headers")
    setColor(worksheet.cell(row, 2), "expense")
    setColor(worksheet.cell(row, 3), "expense")
    setColor(worksheet.cell(row, 4), "expense")
    setColor(worksheet.cell(row, 5), "funds")
    setColor(worksheet.cell(row, 6), "funds")
    setColor(worksheet.cell(row, 7), "funds")
    setColor(worksheet.cell(row, 8), "income")
    setColor(worksheet.cell(row, 9), "income")
    setColor(worksheet.cell(row, 10), "savings")