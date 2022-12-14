from openpyxl import *
from dependencies.excel_functions import *
from dependencies.date import *
import pandas as pd

def addExpenseData(data):
    month = data["Month"] + data["Year"]
    try:
        workbook = load_workbook("./output/Expenditure.xlsx")
    except:
        workbook = createWorkbook(month)
    
    if(month in workbook.sheetnames):
        worksheet = workbook[month]
        worksheet.delete_rows(worksheet.max_row)
    else:
        createSheet(workbook, month)
        worksheet = workbook[month]

    new_row = worksheet.max_row + 1
    worksheet["A" + str(new_row)] = data["Day"] + "-" + getMonthNumber(data["Month"]) + "-" + data["Year"]
    worksheet["B" + str(new_row)] = data["Type"]
    worksheet["C" + str(new_row)] = data["Amount"]
    worksheet["D" + str(new_row)] = data["Mode"]
    setRowColor(worksheet, new_row)

    new_row = new_row + 1
    worksheet["A" + str(new_row)] = "Total"
    worksheet["C" + str(new_row)] = "=SUM(C3:C" + str(new_row - 1) + ")"
    worksheet["F" + str(new_row)] = "=SUM(F3:F" + str(new_row - 1) + ")"
    worksheet["I" + str(new_row)] = "=SUM(I3:I" + str(new_row - 1) + ")"
    worksheet["J" + str(new_row)] = "=I" + str(new_row) + "-F" + str(new_row) + "-C" + str(new_row)
    setRowColor(worksheet, new_row)

    for column in worksheet.columns:
        for cell in column:
            cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')
    workbook._sheets.sort(key = lambda worksheet: worksheet.title[-4:])
    workbook.save("./output/Expenditure.xlsx")

def addFundsData(data):
    month = data["Month"] + data["Year"]
    try:
        workbook = load_workbook("./output/Expenditure.xlsx")
    except:
        workbook = createWorkbook(month)
    
    if(month in workbook.sheetnames):
        worksheet = workbook[month]
        worksheet.delete_rows(worksheet.max_row)
    else:
        createSheet(workbook, month)
        worksheet = workbook[month]

    new_row = worksheet.max_row + 1
    worksheet["A" + str(new_row)] = data["Day"] + "-" + getMonthNumber(data["Month"]) + "-" + data["Year"]
    worksheet["E" + str(new_row)] = "Angel"
    worksheet["F" + str(new_row)] = data["Amount"]
    worksheet["G" + str(new_row)] = data["Mode"]
    setRowColor(worksheet, new_row)

    new_row = new_row + 1
    worksheet["A" + str(new_row)] = "Total"
    worksheet["C" + str(new_row)] = "=SUM(C3:C" + str(new_row - 1) + ")"
    worksheet["F" + str(new_row)] = "=SUM(F3:F" + str(new_row - 1) + ")"
    worksheet["I" + str(new_row)] = "=SUM(I3:I" + str(new_row - 1) + ")"
    worksheet["J" + str(new_row)] = "=I" + str(new_row) + "-F" + str(new_row) + "-C" + str(new_row)
    setRowColor(worksheet, new_row)

    for column in worksheet.columns:
        for cell in column:
            cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')
    workbook._sheets.sort(key = lambda worksheet: worksheet.title[-4:])
    workbook.save("./output/Expenditure.xlsx")

def addIncomeData(data):
    month = data["Month"] + data["Year"]
    try:
        workbook = load_workbook("./output/Expenditure.xlsx")
    except:
        workbook = createWorkbook(month)
    
    if(month in workbook.sheetnames):
        worksheet = workbook[month]
        worksheet.delete_rows(worksheet.max_row)
    else:
        createSheet(workbook, month)
        worksheet = workbook[month]

    new_row = worksheet.max_row + 1
    worksheet["A" + str(new_row)] = data["Day"] + "-" + getMonthNumber(data["Month"]) + "-" + data["Year"]
    worksheet["H" + str(new_row)] = data["Type"]
    worksheet["I" + str(new_row)] = data["Amount"]
    setRowColor(worksheet, new_row)

    new_row = new_row + 1
    worksheet["A" + str(new_row)] = "Total"
    worksheet["C" + str(new_row)] = "=SUM(C3:C" + str(new_row - 1) + ")"
    worksheet["F" + str(new_row)] = "=SUM(F3:F" + str(new_row - 1) + ")"
    worksheet["I" + str(new_row)] = "=SUM(I3:I" + str(new_row - 1) + ")"
    worksheet["J" + str(new_row)] = "=I" + str(new_row) + "-F" + str(new_row) + "-C" + str(new_row)
    setRowColor(worksheet, new_row)

    for column in worksheet.columns:
        for cell in column:
            cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')
    workbook._sheets.sort(key = lambda worksheet: worksheet.title[-4:])
    workbook.save("./output/Expenditure.xlsx")

def addExpenseTypeOrModeData(data):
    try:
        old_data = pd.read_csv("./dependencies/data.csv")
    except:
        old_data = pd.DataFrame()
        print(data)